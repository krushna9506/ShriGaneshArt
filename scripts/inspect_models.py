import openpyxl

path = r'C:\Users\donge\Downloads\Shri_Ganesh_Art_Model_List (1).xlsx'
wb = openpyxl.load_workbook(path, data_only=False)
print('SHEETS', wb.sheetnames)
for ws in wb.worksheets:
    print('\nSHEET', ws.title, 'max_row', ws.max_row, 'max_col', ws.max_column)
    for r in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=False):
        vals = [(c.coordinate, c.value, c.data_type) for c in r]
        print(vals)
