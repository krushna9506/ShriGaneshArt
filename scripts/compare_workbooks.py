import openpyxl
import os

files = [
    r'C:\Users\donge\Downloads\Shri_Ganesh_Art_Model_List.xlsx',
    r'C:\Users\donge\Downloads\Shri_Ganesh_Art_Model_List (1).xlsx',
]
for f in files:
    print('FILE', os.path.basename(f), 'exists', os.path.exists(f))
    wb = openpyxl.load_workbook(f, data_only=False)
    ws = wb.active
    print(' sheets', wb.sheetnames, 'rows', ws.max_row, 'cols', ws.max_column)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=True), start=1):
        print(' row', i, row)
    print('---')
